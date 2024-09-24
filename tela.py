from tkinter import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import xml.etree.ElementTree as ET
import pandas as pd
import unidecode
import unicodedata
import psycopg2
import uuid
from psycopg2 import Error
# nomedoarquivo='exemplo'
anos = []
#sexatas
#departamentos = ['Departamento de Química','Departamento de Física','Departamento de Matemática e Estatística','Departamento de Geociências']
#secate 
#departamentos = []
def criar_planilha_prof(departamentos,professores,doutor,mestre,graduado,colaborador,efetivo,pesquisa,extensao,outros,nacional,internacional,completo,resumo,resumoexp,oripos,oriic,tide,bolsa,financiado,produtividade,colprofessores,coldoutor,colmestre,colgraduado,colcolaborador,colefetivo,colpesquisa,colextensao,coloutros,colnacional,colinternacional,colcompleto,colresumo,colresumoexp,coloripos,coloriic,coltide,colbolsa,colfinanciado,colprodutividade):
        #CRIAÇAO DA PLANILHA
    
    wb = Workbook()
    ws = wb.active

    connection = psycopg2.connect(database="banco2",
                                        host="localhost",
                                        user="postgres",
                                        password="Gumattos2",
                                        port="5432")

    cursor = connection.cursor()
    
    linha = 2
    try:
        for departamento in departamentos:
            print(anos)
            for ano in anos:
                #QUIMICA
                strano = str(ano)
                dep = departamento
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND anosuepg LIKE %s",(dep,'%'+ano+'%'))
                quantprofessor = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND graduacao = 'Doutor'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quantdoutor = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND graduacao = 'Mestre'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quantmestre = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND graduacao = 'Graduado'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quantgraduado = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND status = 'COLABORADOR'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quantcolaborador = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND status = 'EFETIVO'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quantefetivo = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND tide = 'sim'AND anosuepg LIKE %s",(dep,'%'+strano+'%'))
                quanttide = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM projetos WHERE departamento LIKE %s AND tipo = 'PESQUISA' AND anopubli = %s",('%'+dep+'%',strano,))
                quantpesquisa = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM projetos WHERE departamento LIKE %s AND tipo = 'EXTENSAO' AND anopubli = %s",('%'+dep+'%',strano,))
                quantextensao = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM projetos WHERE departamento LIKE %s AND tipo != 'PESQUISA' AND tipo != 'EXTENSAO' AND anopubli = %s",('%'+dep+'%',strano,))
                quantoutros = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM artigos WHERE departamento LIKE %s AND tipo = 'NACIONAL' AND anopubli = %s",('%'+dep+'%',strano,))
                quantnacional = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM artigos WHERE departamento LIKE %s AND tipo = 'INTERNACIONAL' AND anopubli = %s",('%'+dep+'%',strano,))
                quantinternacional = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM congressos WHERE departamento LIKE %s AND tipo = 'COMPLETO' AND anoconclusao = %s",('%'+dep+'%',strano,))
                quantcompleto = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM congressos WHERE departamento LIKE %s AND tipo = 'RESUMO' AND anoconclusao = %s",('%'+dep+'%',strano,))
                quantresumo = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM congressos WHERE departamento LIKE %s AND tipo = 'RESUMO_EXPANDIDO' AND anoconclusao = %s",('%'+dep+'%',strano,))
                quantresumoexp = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE departamento = %s AND tipo = 'DOUTORADO' AND anoconclusao = %s",(dep,strano,))
                quantoridou = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE departamento = %s AND tipo = 'MESTRADO' AND anoconclusao = %s",(dep,strano,))
                quantorimest = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE departamento = %s AND tipo = 'INICIAÇÃO CIENTÍFICA' AND anoconclusao = %s",(dep,strano,))
                quantoriic = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM bolsas WHERE departamento LIKE %s AND ano = %s",('%'+dep+'%',strano,))
                quantbolsa = cursor.fetchone()[0]
                cursor.execute("SELECT SUM(valor) FROM financiados WHERE departamento LIKE %s AND anopubli = %s", ('%' + dep + '%', strano,))
                quantfinanciado = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM produtividade WHERE departamento LIKE %s AND ano = %s",('%'+dep+'%',strano,))
                quantprodutividade = cursor.fetchone()[0]
                strlinha= str(linha)
                ws['A'+strlinha] = ano
                ws['B'+strlinha] = dep
                #COLOCANDO NOME DAS VARIAVEIS NAS COLUNAS
                ws['A1']= 'ANO'
                ws['B1']= 'DEPARTAMENTO'
                ws['C1']= 'SETOR'
                if professores:
                    ws[colprofessores+'1']='PROFESSORES'
                if doutor:
                    ws[coldoutor+'1']='DOUTORES'
                if mestre:
                    ws[colmestre+'1']='MESTRES'
                if graduado:
                    ws[colgraduado+'1']='GRADUADOS'  
                if colaborador:
                    ws[colcolaborador+'1']='COLABORADORES'  
                if efetivo:
                    ws[colefetivo+'1']='EFETIVOS'      
                if pesquisa:
                    ws[colpesquisa+'1']='PROJETOS DE PESQUISA'    
                if extensao:
                    ws[colextensao+'1']='PROJETOS DE EXTENSÃO'      
                if outros:
                    ws[coloutros+'1']='PROJETOS DE OUTRO TIPO'       
                if nacional:
                    ws[colnacional+'1']='ARTIGOS DE CIRCULAÇÃO NACIONAL'      
                if internacional:
                    ws[colinternacional+'1']='ARTIGOS DE CIRCULAÇÃO INTERNACIONAL'      
                if completo:
                    ws[colcompleto+'1']='PUBLICAÇÕES EM CONGRESSOS COMPLETAS'      
                if resumo:
                    ws[colresumo+'1']='PUBLICAÇÕES EM CONGRESSOS RESUMO' 
                if resumoexp:
                    ws[colresumoexp+'1']='PUBLICAÇÕES EM CONGRESSOS RESUMO EXPANDIDO'      
                if oripos:
                    ws[coloripos+'1']='ORIENTAÇÕES DE PÓS-GRADUAÇÃO'
                if oriic:
                    ws[coloriic+'1']='ORIENTAÇÕES DE INICIAÇÃO CIENTÍFICA' 
                if tide:
                    ws[coltide+'1']='TIDE' 
                if bolsa:
                    ws[colbolsa+'1']='BOLSAS DE INICIAÇÃO CIENTÍFICA'
                if financiado:
                    ws[colfinanciado+'1']='VALORES PROJETOS FINANCIADOS'      
                if produtividade:
                    ws[colprodutividade+'1']='BOLSA PRODUTIVIDADE'         
                #ATRIBUIDO VALORES
                if professores:
                    ws[colprofessores+strlinha]=quantprofessor
                if doutor:
                    ws[coldoutor+strlinha]=quantdoutor
                if mestre:
                    ws[colmestre+strlinha]=quantmestre    
                if graduado:
                    ws[colgraduado+strlinha]=quantgraduado       
                if colaborador:
                    ws[colcolaborador+strlinha]=quantcolaborador   
                if efetivo:
                    ws[colefetivo+strlinha]=quantefetivo      
                if pesquisa:
                    ws[colpesquisa+strlinha]=quantpesquisa      
                if extensao:
                    ws[colextensao+strlinha]=quantextensao      
                if outros:
                    ws[coloutros+strlinha]=quantoutros       
                if nacional:
                    ws[colnacional+strlinha]=quantnacional      
                if internacional:
                    ws[colinternacional+strlinha]=quantinternacional      
                if completo:
                    ws[colcompleto+strlinha]=quantcompleto      
                if resumo:
                    ws[colresumo+strlinha]=quantresumo 
                if resumoexp:
                    ws[colresumoexp+strlinha]=quantresumoexp      
                if oripos:
                    ws[coloripos+strlinha]=quantoridou+quantorimest
                if oriic:
                    ws[coloriic+strlinha]=quantoriic  
                if tide:
                    ws[coltide+strlinha]=quanttide
                if bolsa:
                    ws[colbolsa+strlinha]=quantbolsa
                if financiado:
                    ws[colfinanciado+strlinha]=quantfinanciado
                if produtividade:
                    ws[colprodutividade+strlinha]=quantprodutividade    
                if departamento == 'Departamento de Ciências do Solo e Engenharia Agrícola' or departamento =='Departamento de Engenharia Civil'or departamento =='Departamento de Engenharia de Alimentos'or departamento =='Departamento de Engenharia de Materiais'or departamento =='Departamento de Fitotecnia e Fitossanidade/DEFITO'or departamento =='Departamento de Informática'or departamento =='Departamento de Zootecnia':
                    ws['C'+strlinha]='SECATE'
                if departamento =='Departamento de Química'or departamento =='Departamento de Física'or departamento =='Departamento de Matemática e Estatística'or departamento =='Departamento de Geociências':
                    ws['C'+strlinha]='SEXATAS'
                linha+=1

            #FISICA
            #MATEMATICA
            #GEOCIENCIAS
        # Fechar cursor e conexão
        cursor.close()
        connection.close()
        wb.save(entrada.get()+".xlsx")


    except psycopg2.Error as e:
        print("Erro ao executar a consulta SQL:", e)
    finally:
    # Fecha a conexão com o banco de dados
        cursor.close()
        connection.close()
ano_especifico = '2023'
consulta_sql = """
    SELECT p.id_professor, p.nome, COUNT(a.nome) FROM professores p
    LEFT JOIN artigos a ON p.id_professor = a.id_professor
    WHERE a.anopubli = %s
    GROUP BY p.id_professor, p.nome;
"""
connection = psycopg2.connect(database="banco2",
                                    host="localhost",
                                    user="postgres",
                                    password="Gumattos2",
                                    port="5432")

cursor = connection.cursor()

# def artigos(coluna):
#     for depart in departamentos:
#         for ano in anos:
#             cursor.execute("SELECT COUNT(*) FROM professores WHERE departamento = %s AND anopubli = %s",(depart,ano))
#             profs = cursor.fetchone()[0]
def criar_planilha_dep():
    obter_nome_arquivo()
    professores = False
    doutor = False  # Defina as variáveis aqui
    mestre = False
    graduado = False
    colaborador = False
    efetivo = False
    pesquisa = False
    extensao = False
    outros = False
    nacional = False
    internacional = False
    completo = False
    resumo = False
    resumoexp = False
    oripos = False
    oriic = False
    tide= False
    bolsa= False
    financiado = False
    produtividade = False
    colunas = ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W']
    colprofessores = None
    coldoutor = None 
    colmestre = None
    colgraduado = None
    colcolaborador = None
    colefetivo = None
    colpesquisa = None
    colextensao = None
    coloutros = None
    colnacional = None
    colinternacional = None
    colcompleto = None
    colresumo = None
    colresumoexp = None
    coloripos = None
    coloriic = None
    coltide = None
    colbolsa = None
    colfinanciado = None
    colprodutividade = None
    colvariaveis = [colprofessores,coldoutor,colmestre,colgraduado,colcolaborador,colefetivo,colpesquisa,colextensao,coloutros,colnacional,colinternacional,colcompleto,colresumo,colresumoexp,coloripos,coloriic,coltide,colbolsa,colfinanciado,colprodutividade]
    if var_checkboxnome.get():
        professores= True
    if var_checkboxgraduacao.get():
        doutor = True
        mestre = True
        graduado = True
    if var_checkboxstatus.get():
        colaborador = True
        efetivo = True
    if var_checkboxprojetos.get():
        pesquisa = True
        extensao = True
        outros = True
    if var_checkboxartigos.get():
        nacional= True
        internacional = True
    if var_checkboxcongressos.get():
        completo = True
        resumo = True
        resumoexp = True
    if var_checkboxorientacoes.get():
        oripos = True
        oriic = True
    if var_checkboxtide.get():
        tide = True
    if var_checkboxbolsa.get():
        bolsa = True
    if var_checkboxfinanciado.get():
        financiado = True
    if var_checkboxprodutividade.get():
        produtividade = True
    variaveis = [professores,doutor,mestre,graduado,colaborador,efetivo,pesquisa,extensao,outros,nacional,internacional,completo,resumo,resumoexp,oripos,oriic,tide,bolsa,financiado,produtividade]

    aux = 0
    auxcol = 0
    while aux <= 19:
        if variaveis[aux]:
            colvariaveis[aux] = colunas[auxcol]
            auxcol+=1
        aux+=1
    colprofessores = colvariaveis[0]
    coldoutor = colvariaveis[1]
    colmestre = colvariaveis[2]
    colgraduado = colvariaveis[3]
    colcolaborador = colvariaveis[4]
    colefetivo = colvariaveis[5]
    colpesquisa = colvariaveis[6]
    colextensao = colvariaveis[7]
    coloutros = colvariaveis[8]
    colnacional = colvariaveis[9]
    colinternacional = colvariaveis[10]
    colcompleto = colvariaveis[11]
    colresumo = colvariaveis[12]
    colresumoexp = colvariaveis[13]
    coloripos = colvariaveis[14]
    coloriic = colvariaveis[15]
    coltide = colvariaveis[16]
    colbolsa = colvariaveis[17]
    colfinanciado = colvariaveis[18]
    colprodutividade = colvariaveis[19]


    varSEXATAS = False
    varSECATE = False
    #ATRIBUIÇÃO DOS ANOS
    p2019=False
    p2020=False
    p2021=False
    p2022=False
    p2023=False

    if var_checkboxSECATE.get():
        varSECATE = True
    if var_checkboxSEXATAS.get():
        varSEXATAS = True
    if varSECATE is True and varSEXATAS is True:
        departamentos = ['Departamento de Ciências do Solo e Engenharia Agrícola','Departamento de Engenharia Civil','Departamento de Engenharia de Alimentos','Departamento de Engenharia de Materiais','Departamento de Fitotecnia e Fitossanidade/DEFITO','Departamento de Informática','Departamento de Zootecnia','Departamento de Química','Departamento de Física','Departamento de Matemática e Estatística','Departamento de Geociências']
    elif varSECATE:
        departamentos = ['Departamento de Ciências do Solo e Engenharia Agrícola','Departamento de Engenharia Civil','Departamento de Engenharia de Alimentos','Departamento de Engenharia de Materiais','Departamento de Fitotecnia e Fitossanidade/DEFITO','Departamento de Informática','Departamento de Zootecnia']
    elif varSEXATAS:
        departamentos = ['Departamento de Química','Departamento de Física','Departamento de Matemática e Estatística','Departamento de Geociências']
    if var_checkbox2019.get():
        p2019 = True
    if var_checkbox2020.get():
        p2020 = True
    if var_checkbox2021.get():
        p2021 = True
    if var_checkbox2022.get():
        p2022 = True
    if var_checkbox2023.get():
        p2023 = True
    if p2019:
        anos.append('2019')
    if p2020:
        anos.append('2020')
    if p2021:
        anos.append('2021')
    if p2022:
        anos.append('2022')
    if p2023:
        anos.append('2023')
    print (anos)
    # Lógica para atribuição de colunas
    # Exemplo de saída para verificar as colunas atribuídas
    print("Coluna Professores:", colprofessores)
    print("Coluna Doutor:", coldoutor)
    print("Coluna Mestre:", colmestre)
    print("Coluna Graduado:", colgraduado)
    print("Coluna Colaborador:", colcolaborador)
    print("Coluna Efetivo:", colefetivo)
    print("Coluna Pesquisa:", colpesquisa)
    print("Coluna Extensão:", colextensao)
    print("Coluna Outros:", coloutros)
    print("Coluna Nacional:", colnacional)
    print("Coluna Internacional:", colinternacional)
    print("Coluna Completo:", colcompleto)
    print("Coluna Resumo:", colresumo)
    print("Coluna Resumo Expandido:", colresumoexp)
    print("Coluna Orientações Pos:", coloripos)
    print("Coluna Orientações Ic:", coloriic)
    print("Coluna TIDE:", coltide)
    print("Coluna Bolsa:", colbolsa)
    print("Coluna Financiado:", colfinanciado)


    criar_planilha_prof(departamentos,professores,doutor,mestre,graduado,colaborador,efetivo,pesquisa,extensao,outros,nacional,internacional,completo,resumo,resumoexp,oripos,oriic,tide,bolsa,financiado,produtividade,colprofessores,coldoutor,colmestre,colgraduado,colcolaborador,colefetivo,colpesquisa,colextensao,coloutros,colnacional,colinternacional,colcompleto,colresumo,colresumoexp,coloripos,coloriic,coltide,colbolsa,colfinanciado,colprodutividade)
    cursor.close()
    connection.close()
#FUNÇÕES SETOR
def atualizar_SECATE():
    if var_checkboxSECATE.get():
        varSECATE.set(True)
    else:
        varSECATE.set(False)

def atualizar_SEXATAS():
    if var_checkboxSEXATAS.get():
        varSEXATAS.set(True)
    else:
        varSEXATAS.set(False)
#FUNÇÕES DOS ANOS
def atualizar_2019():
    if var_checkbox2019.get():
        var2019.set(True)
    else:
        var2019.set(False)
def atualizar_2020():
    if var_checkbox2020.get():
        var2020.set(True)
    else:
        var2020.set(False)
def atualizar_2021():
    if var_checkbox2021.get():
        var2021.set(True)
    else:
        var2021.set(False)
def atualizar_2022():
    if var_checkbox2022.get():
        var2022.set(True)
    else:
        var2022.set(False)
def atualizar_2023():
    if var_checkbox2023.get():
        var2023.set(True)
    else:
        var2023.set(False)     

        #FUNÇÕES DAS VARIAVEIS   
        
def atualizar_nome():
    if var_checkboxnome.get():
        varnome.set(True)
    else:
        varnome.set(False)
def atualizar_graduacao():
    if var_checkboxgraduacao.get():
        vargraduacao.set(True)
    else:
        vargraduacao.set(False)
def atualizar_departamento():
    if var_checkboxdepartamento.get():
        vardepartamento.set(True)
    else:
        vardepartamento.set(False)
def atualizar_status():
    if var_checkboxstatus.get():
        varstatus.set(True)
    else:
        varstatus.set(False)       
def atualizar_projetos():
    if var_checkboxprojetos.get():
        varprojetos.set(True)
    else:
        varprojetos.set(False)       
def atualizar_artigos():
    if var_checkboxartigos.get():
        varartigos.set(True)
    else:
        varartigos.set(False)    
def atualizar_congressos():
    if var_checkboxcongressos.get():
        varcongressos.set(True)
    else:
        varcongressos.set(False)  
def atualizar_orientacoes():
    if var_checkboxorientacoes.get():
        varorientacoes.set(True)
    else:
        varorientacoes.set(False)     
def atualizar_tide():
    if var_checkboxtide.get():
        vartide.set(True)
    else:
        vartide.set(False)       
def atualizar_bolsa():
    if var_checkboxbolsa.get():
        varbolsa.set(True)
    else:
        varbolsa.set(False)   
def atualizar_financiado():
    if var_checkboxfinanciado.get():
        varfinanciado.set(True)
    else:
        varfinanciado.set(False)     
def atualizar_produtividade():
    if var_checkboxprodutividade.get():
        varprodutividade.set(True)
    else:
        varprodutividade.set(False)             
def obter_nome_arquivo():
    nomedoarquivo = entrada.get()
    # Você pode fazer o que quiser com o nome do arquivo aqui
    print("Nome do arquivo inserido:", nomedoarquivo)
janela = Tk()
janela.title("Criador de Planilhas")
#vars checkbox setor
var_checkboxSECATE = BooleanVar()
var_checkboxSEXATAS = BooleanVar()


#VARS CHECKBOX DOS ANOS
var_checkbox2019 = BooleanVar()
var_checkbox2020 = BooleanVar()
var_checkbox2021 = BooleanVar()
var_checkbox2022 = BooleanVar()
var_checkbox2023 = BooleanVar() 
#VARS CHECKBOX DAS VARIAÁVEIS
var_checkboxnome = BooleanVar()
var_checkboxgraduacao = BooleanVar()
var_checkboxdepartamento = BooleanVar()
var_checkboxstatus = BooleanVar()
var_checkboxprojetos = BooleanVar()
var_checkboxartigos = BooleanVar()
var_checkboxcongressos = BooleanVar()
var_checkboxorientacoes = BooleanVar()
var_checkboxtide = BooleanVar()
var_checkboxbolsa = BooleanVar()
var_checkboxfinanciado = BooleanVar()
var_checkboxprodutividade = BooleanVar()






#VAR SETORES
varSECATE = BooleanVar()
varSECATE.set(False)
varSEXATAS = BooleanVar()
varSEXATAS.set(False)
#VAR DOS ANOS
var2019 = BooleanVar()
var2019.set(False)
var2020 = BooleanVar()
var2020.set(False)
var2021 = BooleanVar()
var2021.set(False)
var2022 = BooleanVar()
var2022.set(False)  
var2023 = BooleanVar()
var2023.set(False)
#VAR DAS VARIAVEIS
varnome = BooleanVar()
varnome.set(False)
vargraduacao = BooleanVar()
vargraduacao.set(False)
vardepartamento = BooleanVar()
vardepartamento.set(False)
varstatus = BooleanVar()
varstatus.set(False)
varprojetos = BooleanVar()
varprojetos.set(False)
varartigos = BooleanVar()
varartigos.set(False)
varcongressos = BooleanVar()
varcongressos.set(False)
varorientacoes = BooleanVar()
varorientacoes.set(False)
vartide = BooleanVar()
vartide.set(False)
varbolsa = BooleanVar()
varbolsa.set(False)
varfinanciado = BooleanVar()
varfinanciado.set(False)
varprodutividade = BooleanVar()
varprodutividade.set(False)
#LABEL DOS ANOS
selecao_ano = Label(janela,text='Selecione o(s) ano(s):')
selecao_ano.grid(column=0,row=0)
#CHECKBOX DOS ANOS
checkbox2019 = Checkbutton(janela, text="2019", variable=var_checkbox2019, command=atualizar_2019)
checkbox2019.grid(row=2, column=0)
checkbox2020 = Checkbutton(janela, text="2020", variable=var_checkbox2020, command=atualizar_2020)
checkbox2020.grid(row=2, column=1, padx=20)
checkbox2021 = Checkbutton(janela, text="2021", variable=var_checkbox2021, command=atualizar_2021)
checkbox2021.grid(row=2, column=2, padx=20)
checkbox2022 = Checkbutton(janela, text="2022", variable=var_checkbox2022, command=atualizar_2022)
checkbox2022.grid(row=2, column=3, padx=20)
checkbox2023 = Checkbutton(janela, text="2023", variable=var_checkbox2023, command=atualizar_2023)
checkbox2023.grid(row=2, column=4, padx=20)
#LABEL DAS VARIAVEIS
selecao_variaveis = Label(janela,text='Selecione a(s) variaveis(s):')
selecao_variaveis.grid(column=0,row=3)
#CHECKBOX DAS VARIAVEIS
checkboxnome = Checkbutton(janela, text="QUANTIDADEPROFS", variable=var_checkboxnome, command=atualizar_nome)
checkboxnome.grid(row=4, column=0, padx=20)
checkboxgraduacao = Checkbutton(janela, text="GRADUAÇÃO", variable=var_checkboxgraduacao, command=atualizar_graduacao)
checkboxgraduacao.grid(row=4, column=1, padx=20)
# checkboxdepartamento = Checkbutton(janela, text="Departamento", variable=var_checkboxdepartamento, command=atualizar_departamento)
# checkboxdepartamento.grid(row=4, column=2, padx=20)
checkboxstatus = Checkbutton(janela, text="STATUS", variable=var_checkboxstatus, command=atualizar_status)
checkboxstatus.grid(row=4, column=2, padx=20)
checkboxprojetos = Checkbutton(janela, text="PROJETOS", variable=var_checkboxprojetos, command=atualizar_projetos)
checkboxprojetos.grid(row=4, column=3, padx=20)
checkboxartigos = Checkbutton(janela, text="ARTIGOS", variable=var_checkboxartigos, command=atualizar_artigos)
checkboxartigos.grid(row=4, column=4, padx=20)
checkboxcongressos = Checkbutton(janela, text="CONGRESSOS", variable=var_checkboxcongressos, command=atualizar_congressos)
checkboxcongressos.grid(row=4, column=5, padx=20)
checkboxorientacoes = Checkbutton(janela, text="ORIENTAÇÕES", variable=var_checkboxorientacoes, command=atualizar_orientacoes)
checkboxorientacoes.grid(row=4, column=6, padx=20)
checkboxtide = Checkbutton(janela, text="TIDE", variable=var_checkboxtide, command=atualizar_tide)
checkboxtide.grid(row=4, column=7, padx=20)
checkboxbolsa = Checkbutton(janela, text="BOLSA", variable=var_checkboxbolsa, command=atualizar_bolsa)
checkboxbolsa.grid(row=4, column=8, padx=20)
checkboxfinanciado = Checkbutton(janela, text="FINANCIADOS", variable=var_checkboxfinanciado, command=atualizar_financiado)
checkboxfinanciado.grid(row=4, column=9, padx=20)
checkboxprodutividade = Checkbutton(janela, text="PRODUTIVIDADE", variable=var_checkboxprodutividade, command=atualizar_produtividade)
checkboxprodutividade.grid(row=4, column=10, padx=20)
selecao_setor = Label(janela,text='Selecione o(s) setor(es):')
selecao_setor.grid(column=0,row=5)
checkboxSECATE = Checkbutton(janela, text="SECATE", variable=var_checkboxSECATE, command=atualizar_SECATE)
checkboxSECATE.grid(row=6, column=0)
checkboxSEXATAS = Checkbutton(janela, text="SEXATAS", variable=var_checkboxSEXATAS, command=atualizar_SEXATAS)
checkboxSEXATAS.grid(row=6, column=1)
#NOME DO ARQUIVO
nomearq = Label(janela,text='Nome do arquivo:')
nomearq.grid(column=3,row=8)
entrada = Entry(janela)
entrada.grid(row=8, column=4)
botao_criar = Button(janela, text="Criar Planilha",command=criar_planilha_dep)
botao_criar.grid(row=9, column=3, columnspan=3, pady=10)
janela.mainloop()
