import os
import unicodedata
from openpyxl import load_workbook

def remover_acentos(txt):
    if txt is None:
        return ""
    return ''.join(ch for ch in unicodedata.normalize('NFD', str(txt)) if not unicodedata.combining(ch))

def ler_planilhas(pasta, nome):
    anosUEPG = []
    for arquivo in os.listdir(pasta):
        if arquivo.endswith('.xlsx'):
            caminho_arquivo = os.path.join(pasta, arquivo)
            planilha = load_workbook(caminho_arquivo)
            for nome_planilha in planilha.sheetnames:
                planilha_ativa = planilha[nome_planilha]
                for linha in planilha_ativa.iter_rows(min_row=3, min_col=1, max_col=3, values_only=True):
                    ano, _, nome_planilha = linha
                    nome_sem_acentos = remover_acentos(nome_planilha).lower()
                    if nome_sem_acentos == nome.lower():
                        anosUEPG.append(str(ano))
    return anosUEPG

# Defina o nome que você quer procurar
nome = 'Jaime Alberti Gomes'

# Diretório do projeto
diretorio_projeto = os.path.dirname(os.path.abspath(__file__))

# Defina as pastas onde estão as planilhas
pasta_colaboradores = os.path.join(diretorio_projeto, 'COLABORADORES')
pasta_efetivos = os.path.join(diretorio_projeto, 'EFETIVOS')

anos_colaboradores = ler_planilhas(pasta_colaboradores, nome)
anos_efetivos = ler_planilhas(pasta_efetivos, nome)

anosUEPG = anos_colaboradores + anos_efetivos
anosUEPG_concatenados = ', '.join(anosUEPG)

print(f"Anos da UEPG para '{nome}': {anosUEPG_concatenados}")
