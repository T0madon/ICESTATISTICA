import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import xml.etree.ElementTree as ET
import pandas as pd
import unidecode
import unicodedata
import psycopg2
import uuid
from psycopg2 import Error
import os

try:
    # Conectando ao banco de dados
    connection = psycopg2.connect(database="banco2",
                                        host="localhost",
                                        user="postgres",
                                        password="Gumattos2",
                                        port="5432")

    cursor = connection.cursor()

    varformacao = 'Graduado(a)'
    contador = 0
    contplanilha = 1
    while contador < 10:
        contador += 1
        idprofessor = str(uuid.uuid4())
        strcontador = str(contador)
        tree = ET.parse('./xmls/'+strcontador+'.xml')
        root = tree.getroot()
        # busca nome e pais de nascimento
        for gerais in root.iter('DADOS-GERAIS'):
            nome1 = gerais.attrib['NOME-COMPLETO']
            nome = unidecode.unidecode(nome1)
        for formacao in root.iter('FORMACAO-ACADEMICA-TITULACAO'):
            doutor = formacao.find('DOUTORADO')
            for doutor in root.iter('DOUTORADO'):
                andamentodoutorado = doutor.attrib['STATUS-DO-CURSO']

            mestre = formacao.find('MESTRADO')
            for mestre in root.iter('MESTRADO'):
                andamentomestrado = mestre.attrib['STATUS-DO-CURSO']

            graduado = formacao.find('GRADUACAO')
            if doutor != None and andamentodoutorado != 'EM_ANDAMENTO':
                varformacao = 'Doutor'
            elif mestre != None and andamentomestrado != 'EM_ANDAMENTO':
                varformacao = 'Mestre'

            elif graduado != None:
                varformacao = 'Graduado'
                        #departamento
        nomedepart = '0'
        nomereal = 'a'
        arquivodepartefetivo = openpyxl.load_workbook(
            'Efetivos_por_departamento_e_ano.xlsx')
        atualplan = arquivodepartefetivo['Planilha1']
        contdepart = 1
        while contdepart < 688:
            strcontdepart = str(contdepart)
            nomereal1 = (atualplan['C' + strcontdepart].value)
            if nomereal1 is not None:
                nomereal = unidecode.unidecode(nomereal1)
            if nomereal == nome:
                nomedepart = (atualplan['B' + strcontdepart].value)
            contdepart += 1
        if nomedepart == '0':
            arquivodepartcolab = openpyxl.load_workbook(
                'Colaborador_por_departamento_e_ano.xlsx')
            atualplan = arquivodepartcolab['Planilha1']
            contdepart = 1
            nomereal = 'a'
            while contdepart < 300:
                strcontdepart = str(contdepart)
                nomereal1 = (atualplan['C' + strcontdepart].value)
                if nomereal1 is not None:
                    nomereal = unidecode.unidecode(nomereal1)
                if nomereal == nome:
                    nomedepart = (
                        atualplan['B' + strcontdepart].value)
                contdepart += 1        
        #status
        theFile = openpyxl.load_workbook(
                    'Colaborador_por_departamento_e_ano.xlsx')
        atualplan = theFile['Planilha1']
        cont = 1
        nomereal = 'a'
        varcolaborador = 0
        varefetivo = 0
        while cont < 300:
            strcont = str(cont)
            nomereal1 = (atualplan['C' + strcont].value)
            if nomereal1 is not None:
                nomereal = unidecode.unidecode(nomereal1)
            if nomereal == nome:
                varcolaborador = 1
                status= 'Colaborador'

            cont += 1
        if varcolaborador == 0:
            varefetivo = 1
            status= 'Efetivo'

        # Executar a consulta
   
        cursor.execute("SELECT COUNT(*) FROM professores WHERE nome = %s", (nome,))
        total = cursor.fetchone()[0]

        # Verificar se o valor existe na tabela
        if total > 0:
            print("O nome existe na tabela 'professores'."+(nome))
            cursor.execute("SELECT id_professor FROM professores WHERE nome= %s", (nome,))
            idprofessor = cursor.fetchone()[0]
        else:
            cursor.execute('''INSERT INTO public.professores (id_professor, nome, graduacao, departamento, status) 
                  VALUES (%s, %s, %s, %s, %s);''', 
                  (idprofessor, nome, varformacao, nomedepart, status))
            connection.commit()
        
        #INSERINDO PROJETOS NO BANCO
            
        anoprojeto = 0

        for participas in root.iter('ATIVIDADES-DE-PARTICIPACAO-EM-PROJETO'):
            part = participas.find('PARTICIPACAO-EM-PROJETO')

        for part in root.iter('PARTICIPACAO-EM-PROJETO'):
            pesquisa = part.find('PROJETO-DE-PESQUISA')

        for titulo in root.iter('PROJETO-DE-PESQUISA'):
            anoprojeto = titulo.attrib['ANO-INICIO']
            natureza = titulo.attrib['NATUREZA']
            tituloprojeto1 = titulo.attrib['NOME-DO-PROJETO']
            if tituloprojeto1 is not None:
                tituloprojeto2 = unidecode.unidecode(tituloprojeto1) 
                tituloprojeto3 = tituloprojeto2.upper()
                tituloprojeto = tituloprojeto3.replace("'", "")
            cursor.execute("SELECT COUNT(*) FROM projetos WHERE nome = %s", (tituloprojeto,))
            tituloprojetocomparar = cursor.fetchone()[0]
            # Verificar se o valor existe na tabela
            if tituloprojetocomparar >0:
                cursor.execute("SELECT id_professor FROM projetos WHERE nome = %s", (tituloprojeto,))
                id_professorcomparar = cursor.fetchone()[0]
                arrayid_professorcomparar = id_professorcomparar.split(",")
                # Verificar se o valor existe na tabela
                if idprofessor not in arrayid_professorcomparar :
                    cursor.execute("""      UPDATE projetos
                                            SET id_professor = CONCAT(id_professor, ',',%s)
                                            WHERE nome = %s;
                                        """,(idprofessor,tituloprojeto))
                    connection.commit()
                    cursor.execute("SELECT departamento FROM projetos WHERE nome= %s", (tituloprojeto,))
                    departcomparar = cursor.fetchone()[0]
                    arraydepartcomparar = departcomparar.split(',')
                    if nomedepart not in arraydepartcomparar:
                        print("O nome do projeto existe na tabela 'projetos'."+(tituloprojeto))
                        cursor.execute("""  UPDATE projetos
                                            SET departamento = CONCAT(departamento, ','%s)
                                            WHERE nome = %s;
                                        """,(nomedepart,idprofessor,tituloprojeto))
                        connection.commit()
            else:
                cursor.execute('''INSERT into public.projetos (id_professor,nome,anopubli,tipo,departamento) 
                values('%s','%s','%s','%s','%s');
                    ''' % ((idprofessor),(tituloprojeto),(anoprojeto),(natureza),(nomedepart)))
                connection.commit()

        #INSERINDO ARTIGOS NO BANCO
        anoartigo = 0
        nacional = 0
        internacional = 0
        tipoartigo = ''
        tituloartigo=''
        for artigos1 in root.iter('ARTIGOS-PUBLICADOS'):
            artigo1 = artigos1.find('ARTIGO-PUBLICADO')

        for artigos2 in root.iter('ARTIGO-PUBLICADO'):
            artigo2 = artigos2.find('DADOS-BASICOS-DO-ARTIGO')

        for artigo3 in root.iter('DADOS-BASICOS-DO-ARTIGO'):
            anoartigo = artigo3.attrib['ANO-DO-ARTIGO']
            idioma = artigo3.attrib['IDIOMA']
            tituloartigo1 = artigo3.attrib['TITULO-DO-ARTIGO']
            if tituloartigo1 is not None:
                tituloartigo2 = unidecode.unidecode(tituloartigo1) 
                tituloartigo3 = tituloartigo2.upper()
                tituloartigo = tituloartigo3.replace("'", "")
            if idioma == 'Português':
                tipoartigo='NACIONAL'
            else:
                tipoartigo='INTERNACIONAL'

            cursor.execute("SELECT COUNT(*) FROM artigos WHERE nome = %s", (tituloartigo,))
            tituloartigocomparar = cursor.fetchone()[0]
            if tituloartigocomparar > 0:
                cursor.execute("SELECT id_professor FROM artigos WHERE nome = %s", (tituloartigo,))
                id_professor_atual = cursor.fetchone()[0]
                arrayidprofessoratual = id_professor_atual.split(",")
                if  idprofessor not in arrayidprofessoratual:
                    cursor.execute("""      UPDATE artigos
                                            SET id_professor = CONCAT(id_professor, ',',%s)
                                            WHERE nome = %s;
                                        """,(idprofessor,tituloartigo))
                    connection.commit()
                    cursor.execute("SELECT departamento FROM artigos WHERE nome= %s", (tituloartigo,))
                    departcomparar = cursor.fetchone()[0]
                    arraydepartcomparar = departcomparar.split(",")
                    if nomedepart not in arraydepartcomparar :
                        print("O nome do artigo existe na tabela 'artigos'."+(tituloartigo))
                        cursor.execute("""  UPDATE artigos
                                            SET departamento = CONCAT(departamento, ',', %s),
                                            WHERE nome = %s;
                                        """,(nomedepart,tituloartigo))
                        connection.commit()
            else:
                cursor.execute('''INSERT into public.artigos (id_professor,nome,anopubli,tipo,departamento) 
                values('%s','%s','%s','%s','%s');
                    ''' % ((idprofessor),(tituloartigo),(anoartigo),(tipoartigo),(nomedepart)))
                connection.commit()
#          # ORIENTAÇÕES   
        anoori = 0
        for prod in root.iter('OUTRA-PRODUCAO'):
            prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

        for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
            ori1 = ori.find('ORIENTACOES-CONCLUIDAS-PARA-MESTRADO')

        for ori1 in root.iter('ORIENTACOES-CONCLUIDAS-PARA-MESTRADO'):
            mestreano = ori1.find(
                'DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO')
        for mestreano in root.iter('DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO'):
            anodomestrado = mestreano.attrib['ANO']
            tituloorimestrado1 = mestreano.attrib['TITULO']
            if tituloorimestrado1 is not None:
                tituloorimestrado2 = unidecode.unidecode(tituloorimestrado1)
                tituloorimestrado3 = tituloorimestrado2.upper()
                tituloorimestrado = tituloorimestrado3.replace("'", "")
                tipoori = 'MESTRADO'
            cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (tituloorimestrado,)),
            tituloorimestradocomparar = cursor.fetchone()[0]
            if tituloorimestradocomparar > 0:
                cursor.execute("SELECT id_professor FROM orientacoes WHERE nome = %s", (tituloorimestrado,))
                id_professor_atual = cursor.fetchone()[0]
                arrayidprofessoratual = id_professor_atual.split(",")
                if  idprofessor not in arrayidprofessoratual:
                    cursor.execute("""      UPDATE orientacoes
                                            SET id_professor = CONCAT(id_professor, ',',%s)
                                            WHERE nome = %s;
                                        """,(idprofessor,tituloorimestrado))
                    connection.commit()
                    cursor.execute("SELECT departamento FROM orientacoes WHERE nome= %s", (tituloorimestrado,))
                    id_professor_atual = cursor.fetchone()[0]
                    departcomparar = cursor.fetchone()[0]
                    arraydepartcomparar = departcomparar.split(",")
                    if nomedepart not in arraydepartcomparar :
                        print("O nome da orientacao existe na tabela 'orientacoes'."+(tituloorimestrado))
                        cursor.execute("""  UPDATE orientacoes
                                            SET departamento = CONCAT(departamento, ',', %s),
                                            WHERE nome = %s;
                                        """,(nomedepart,tituloorimestrado))
                        connection.commit()
            else:
                cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                values('%s','%s','%s','%s','%s');
                    ''' % ((idprofessor),(tituloorimestrado),(anodomestrado),(tipoori),(nomedepart)))
                connection.commit()
# # DOUTORADO
        for prod in root.iter('OUTRA-PRODUCAO'):
            prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

        for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
            ori1 = ori.find('ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO')

        for ori1 in root.iter('ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO'):
            doutorano = ori1.find(
                'DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO')
        for doutorano in root.iter('DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO'):
            anododoutorado = doutorano.attrib['ANO']
            titulooridoutorado1 = doutorano.attrib['TITULO']
            if titulooridoutorado1 is not None:
                titulooridoutorado2 = unidecode.unidecode(titulooridoutorado1)
                titulooridoutorado3 = titulooridoutorado2.upper()
                titulooridoutorado = titulooridoutorado3.replace("'", "")
                tipoori = 'DOUTORADO'
            cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (titulooridoutorado,)),
            titulooridoutoradocomparar = cursor.fetchone()[0]
            # Verificar se o valor existe na tabela
            if titulooridoutoradocomparar > 0:
                cursor.execute("SELECT id_professor FROM orientacoes WHERE nome = %s", (titulooridoutorado,))
                id_professor_atual = cursor.fetchone()[0]
                arrayidprofessoratual = id_professor_atual.split(",")
                if  idprofessor not in arrayidprofessoratual:
                    cursor.execute("""      UPDATE orientacoes
                                            SET id_professor = CONCAT(id_professor, ',',%s)
                                            WHERE nome = %s;
                                        """,(idprofessor,titulooridoutorado))
                    connection.commit()
                    cursor.execute("SELECT departamento FROM orientacoes WHERE nome= %s", (titulooridoutorado,))
                    id_professor_atual = cursor.fetchone()[0]
                    departcomparar = cursor.fetchone()[0]
                    arraydepartcomparar = departcomparar.split(",")
                    if nomedepart not in arraydepartcomparar :
                        print("O nome da orientacao existe na tabela 'orientacoes'."+(titulooridoutorado))
                        cursor.execute("""  UPDATE orientacoes
                                            SET departamento = CONCAT(departamento, ',', %s),
                                            WHERE nome = %s;
                                        """,(nomedepart,titulooridoutorado))
                        connection.commit()
            else:
                cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                values('%s','%s','%s','%s','%s');
                    ''' % ((idprofessor),(titulooridoutorado),(anododoutorado),(tipoori),(nomedepart)))
                connection.commit()

# # IC
        for prod in root.iter('OUTRA-PRODUCAO'):
            prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

        for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
            ori1 = ori.find('OUTRAS-ORIENTACOES-CONCLUIDAS')

        for ori1 in root.iter('OUTRAS-ORIENTACOES-CONCLUIDAS'):
            icano = ori1.find(
                'DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS')
        for icano in root.iter('DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS'):
            anodoic = icano.attrib['ANO']
            deic = icano.attrib['NATUREZA']
            titulooriic1 = icano.attrib['TITULO']
            if deic == 'INICIACAO_CIENTIFICA':
                if titulooriic1 is not None:
                    titulooriic2 = unidecode.unidecode(titulooriic1)
                    titulooriic3 = titulooriic2.upper()
                    titulooriic = titulooriic3.replace("'", "")
                    tipoori = 'INICIAÇÃO CIENTÍFICA'
                    cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (titulooriic,)),
                    titulooriiccomparar = cursor.fetchone()[0]
                    # Verificar se o valor existe na tabela
                    if titulooriiccomparar > 0:
                        cursor.execute("SELECT id_professor FROM orientacoes WHERE nome = %s", (titulooriic,))
                        id_professor_atual = cursor.fetchone()[0]
                        arrayidprofessoratual = id_professor_atual.split(",")
                        if  idprofessor not in arrayidprofessoratual:
                            cursor.execute("""      UPDATE orientacoes
                                                    SET id_professor = CONCAT(id_professor, ',',%s)
                                                    WHERE nome = %s;
                                                """,(idprofessor,titulooriic))
                            connection.commit()
                            cursor.execute("SELECT departamento FROM orientacoes WHERE nome= %s", (titulooriic,))
                            departcomparar = cursor.fetchone()[0]
                            arraydepartcomparar = departcomparar.split(",")
                            if nomedepart not in arraydepartcomparar :
                                print("O nome da orientacao existe na tabela 'orientacoes'."+(titulooriic))
                                cursor.execute("""  UPDATE orientacoes
                                                    SET departamento = CONCAT(departamento, ',', %s),
                                                    WHERE nome = %s;
                                                """,(nomedepart,titulooriic))
                                connection.commit()
                    else:
                        cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                        values('%s','%s','%s','%s','%s');
                            ''' % ((idprofessor),(titulooriic),(anodoic),(tipoori),(nomedepart)))
                        connection.commit()
        # PUBLICACOES EM CONGRESSO
        for biblio in root.iter('PRODUCAO-BIBLIOGRAFICA'):
            eventos = biblio.find('TRABALHOS-EM-EVENTOS')

        for eventos2 in root.iter('TRABALHOS-EM-EVENTOS'):
            evento = eventos2.find('TRABALHO-EM-EVENTOS')

        for evento2 in root.iter('TRABALHO-EM-EVENTOS'):
            dadoscongre = evento2.find('DADOS-BASICOS-DO-TRABALHO')
            tipocongre = dadoscongre.attrib['NATUREZA']
            anocongre = dadoscongre.attrib['ANO-DO-TRABALHO']
            titulocongre1 = dadoscongre.attrib['TITULO-DO-TRABALHO']
            if titulocongre1 is not None:
                    titulocongre2 = unidecode.unidecode(titulocongre1)
                    titulocongre3 = titulocongre2.upper()
                    titulocongre = titulocongre3.replace("'", "")
                    cursor.execute("SELECT COUNT(*) FROM congressos WHERE nome = %s", (titulocongre,)),
                    titulocongrecomparar = cursor.fetchone()[0]
                    # Verificar se o valor existe na tabela
                    if titulocongrecomparar > 0:
                        cursor.execute("SELECT id_professor FROM congressos WHERE nome = %s", (titulocongre,))
                        id_professor_atual = cursor.fetchone()[0]
                        arrayidprofessoratual = id_professor_atual.split(",")
                        if  idprofessor not in arrayidprofessoratual:
                            cursor.execute("""      UPDATE congressos
                                                    SET id_professor = CONCAT(id_professor, ',',%s)
                                                    WHERE nome = %s;
                                                """,(idprofessor,titulocongre))
                            connection.commit()
                            cursor.execute("SELECT departamento FROM congressos WHERE nome= %s", (titulocongre,))
                            departcomparar = cursor.fetchone()[0]
                            arraydepartcomparar = departcomparar.split(",")
                            if nomedepart not in arraydepartcomparar :
                                print("O nome da publicaão existe na tabela 'congressos'."+(titulocongre))
                                cursor.execute("""  UPDATE congressos
                                                    SET departamento = CONCAT(departamento, ',', %s),
                                                    WHERE nome = %s;
                                                """,(nomedepart,titulocongre))
                                connection.commit()
        
                    else:
                        cursor.execute('''INSERT into public.congressos (id_professor,nome,anoconclusao,tipo,departamento) 
                        values('%s','%s','%s','%s','%s');
                            ''' % ((idprofessor),(titulocongre),(anocongre),(tipocongre),(nomedepart)))
                        connection.commit()
        
        # Função para comparar os nomes ignorando os acentos
        def comparar_nomes(nome_a, nome_b):
            return unidecode.unidecode(str(nome_a).lower()) == unidecode.unidecode(str(nome_b).lower())

        # Leitura das planilhas
        colaboradores_df = pd.read_excel("COLABORADORES_SEXATAS.xlsx")
        efetivos_df = pd.read_excel("Efetivos_SEXATAS.xlsx")
        tide_df = pd.read_excel("TIDE_SEXATAS.xlsx")

        # Extrair os valores da coluna C, ignorando os valores que não são strings
        colaboradores_nomes = [str(nome) for nome in colaboradores_df.iloc[:, 2].tolist() if isinstance(nome, str)]
        efetivos_nomes = [str(nome) for nome in efetivos_df.iloc[:, 2].tolist() if isinstance(nome, str)]
        tide_nomes = [str(nome) for nome in tide_df.iloc[:, 2].tolist() if isinstance(nome, str)]

        # Verificar se o nome está presente nas listas de nomes
        status = ''
        tide = ''

        if any(comparar_nomes(nome, n) for n in colaboradores_nomes):
            status = 'COLABORADOR'
        elif any(comparar_nomes(nome, n) for n in efetivos_nomes):
            status = 'EFETIVO'
        else:
            status = 'DESCONHECIDO'

        if any(comparar_nomes(nome, n) for n in tide_nomes):
            tide = 'sim'
        else:
            tide = 'nao'

        print("Status para", nome, ":", status)
        print("Tide para", nome, ":", tide)
        status = status[:12]
        cursor.execute("""  
            UPDATE professores
            SET status = %s
            WHERE nome = %s;
        """, (status, nome))
        connection.commit()
        tide = tide[:12]
        cursor.execute("""  
            UPDATE professores
            SET tide = %s
            WHERE nome = %s;
        """, (tide, nome))
        connection.commit()


        def percorrer_planilhas():
            anosUEPG = ''  # Inicialize a variável anosUEPG
            for filename in os.listdir("PLANILHAS"):
                if filename.endswith(".xlsx"):
                    path = os.path.join("PLANILHAS", filename)
                    df = pd.read_excel(path)
                    nomes = [str(nome) for nome in df.iloc[:, 2].tolist() if isinstance(nome, str)]
                    anos = df.iloc[:, 0].tolist()
                    for i, n in enumerate(nomes):
                        if comparar_nomes(nome, n):
                            if anosUEPG == '':
                                anosUEPG = str(anos[i])
                            else:
                                anosUEPG += ',' + str(anos[i])
            return anosUEPG

        # Chamada da função para percorrer as planilhas
        anosUEPG = percorrer_planilhas()
        print("Anos da UEPG para", nome, ":", anosUEPG)

        cursor.execute("""  
            UPDATE professores
            SET anosUEPG = %s
            WHERE nome = %s;
        """, (anosUEPG, nome))
        connection.commit()

        #PLANILHAS BOLSAs

        # Obtém o diretório atual onde está o código
        diretorio_atual = os.path.dirname(os.path.abspath(__file__))

        # Pasta onde estão os dados
        pasta_bolsasic = "BOLSASIC"

        # Percorre os anos dentro da pasta BOLSASIC
        for ano in range(2019, 2025):  # Modifique conforme os anos desejados
            pasta_ano = os.path.join(diretorio_atual, pasta_bolsasic, str(ano))
            if os.path.exists(pasta_ano):
                # Lista os arquivos na pasta do ano
                arquivos = os.listdir(pasta_ano)
                # Itera sobre os arquivos
                for arquivo in arquivos:
                    # Verifica se o arquivo é uma planilha (pode adaptar essa verificação conforme o tipo de arquivo)
                    if arquivo.endswith('.xlsx') or arquivo.endswith('.xls'):
                        # Lê a planilha
                        caminho_arquivo = os.path.join(pasta_ano, arquivo)
                        
                        # Lê a planilha, especificando que o cabeçalho está na linha 8 (0-indexed)
                        df = pd.read_excel(caminho_arquivo, header=8)
                        
                        # Remove os acentos dos nomes na coluna 'ORIENTADOR(A)' e 'SUBPROJETO DO' da planilha
                        df['ORIENTADOR(A)'] = df['ORIENTADOR(A)'].astype(str).apply(unidecode)
                        df['SUBPROJETO DO'] = df['SUBPROJETO DO'].astype(str).apply(unidecode)
                        
                        # Filtra as linhas onde o nome do professor corresponde ao nome desejado
                        df_filtrado = df[df['ORIENTADOR(A)'] == unidecode(nome)]
                        
                        # Itera sobre as linhas filtradas
                        for index, row in df_filtrado.iterrows():
                            # Imprime o nome do trabalho e o ano da pasta
                            print("Professor:", row['ORIENTADOR(A)'])
                            print("Trabalho:", row['SUBPROJETO DO'])
                            print("Ano do trabalho:", ano)
                            titulobolsa = row['SUBPROJETO DO']
                            anobolsa = ano
                            print()  # Adiciona uma linha em branco entre as impressões
                            cursor.execute("SELECT COUNT(*) FROM bolsas WHERE nome = %s", (titulobolsa,))
                            titulobolsacomparar = cursor.fetchone()[0]
                            if titulobolsacomparar < 1:
                                cursor.execute('''INSERT into public.bolsas (id_professor,nome,ano,departamento) 
                                values('%s','%s','%s','%s','%s');
                            ''' % ((idprofessor),(titulobolsa),(anobolsa),(nomedepart)))
                                connection.commit()

except (Exception, psycopg2.Error) as error:
    print("Erro ao conectar ao PostgreSQL:", error)
finally:
# Fechar a conexão
    if connection:
        cursor.close()
        connection.close()