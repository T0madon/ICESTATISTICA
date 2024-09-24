import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import xml.etree.ElementTree as ET
import pandas as pd
import unidecode
import unicodedata
import psycopg2
import uuid
from psycopg2 import Error
import os
from datetime import datetime


try:
    # Conectando ao banco de dados
    connection = psycopg2.connect(database="banco2",
                                        host="localhost",
                                        user="postgres",
                                        password="Gumattos2",
                                        port="5432")

    cursor = connection.cursor()

    varformacao = 'Graduado(a)'
    # Diretório onde os arquivos XML estão localizados
    directory = './xmls'

    # Lista todos os arquivos no diretório e filtra apenas os arquivos XML
    xml_files = [f for f in os.listdir(directory) if f.endswith('.xml')]
    nome = 'None'
    # Itera sobre cada arquivo XML encontrado
    for xml_file in xml_files:
        idprofessor = str(uuid.uuid4())
        file_path = os.path.join(directory, xml_file)
        
        # Carrega o arquivo XML
        tree = ET.parse(file_path)
        root = tree.getroot()
        # busca nome e pais de nascimento
        print(nome)
        for gerais in root.iter('DADOS-GERAIS'):
            nome1 = gerais.attrib['NOME-COMPLETO']
            nome = unidecode.unidecode(nome1)
            print(nome)
        for formacao in root.iter('FORMACAO-ACADEMICA-TITULACAO'):
            doutor = formacao.find('DOUTORADO')
            for doutor in root.iter('DOUTORADO'):
                andamentodoutorado = doutor.attrib['STATUS-DO-CURSO']

            mestre = formacao.find('MESTRADO')
            for mestre in root.iter('MESTRADO'):
                andamentomestrado = mestre.attrib['STATUS-DO-CURSO']

            graduado = formacao.find('GRADUACAO')
            if doutor != None and andamentodoutorado != 'EM_ANDAMENTO':
                varformacao = 'Doutor'
            elif mestre != None and andamentomestrado != 'EM_ANDAMENTO':
                varformacao = 'Mestre'

            elif graduado != None:
                varformacao = 'Graduado'
                        #departamento
        def remove_accents(text):
            if isinstance(text, str):
                return unidecode.unidecode(text)
            else:
                return text

        def find_department_and_status(nome, efetivos, colaboradores):
            for df, status in [(efetivos, 'EFETIVO'), (colaboradores, 'COLABORADOR')]:
                df['nm_pessoa_completo'] = df['nm_pessoa_completo'].apply(remove_accents)
                if nome in df['nm_pessoa_completo'].values:
                    departamento = df.loc[df['nm_pessoa_completo'] == nome, 'DEPARTAMENTO'].iloc[0]
                    return departamento, status
            return None, None

        # Diretório dos arquivos EFETIVOS e COLABORADORES
        diretorio = os.path.dirname(__file__)
        caminho_efetivos = os.path.join(diretorio, 'EFETIVOS')
        caminho_colaboradores = os.path.join(diretorio, 'COLABORADORES')

        efetivos = []
        colaboradores = []

        for arquivo in os.listdir(caminho_efetivos):
            if arquivo.endswith('.xlsx'):
                efetivos.append(pd.read_excel(os.path.join(caminho_efetivos, arquivo), header=2))  # Começa a leitura a partir da terceira linha (índice 2)

        for arquivo in os.listdir(caminho_colaboradores):
            if arquivo.endswith('.xlsx'):
                colaboradores.append(pd.read_excel(os.path.join(caminho_colaboradores, arquivo), header=2))  # Começa a leitura a partir da terceira linha (índice 2)

        # Concatenação dos dataframes
        efetivos = pd.concat(efetivos)
        colaboradores = pd.concat(colaboradores)

        # Remoção de acentos e sinais dos nomes
        efetivos['nm_pessoa_completo'] = efetivos['nm_pessoa_completo'].apply(remove_accents)
        colaboradores['nm_pessoa_completo'] = colaboradores['nm_pessoa_completo'].apply(remove_accents)


        # Procura pelo departamento do nome inserido
        departamento, status = find_department_and_status(nome, efetivos, colaboradores)
        nomedepart=departamento
        if nomedepart is None:
            print('Professor não encontrado nas planilhas'+ nome)
        # if departamento:
        #     print(f"O departamento de {nome} é: {departamento}")
        #     print(f"Status: {status}")
        # else:
        #     print("Nome não encontrado nas planilhas.")

        
        #SETOR
        setor = 'DESC'
        if nomedepart =='Departamento de Química' or  nomedepart == 'Departamento de Física' or  nomedepart == 'Departamento de Matemática e Estatística' or  nomedepart == 'Departamento de Geociências':
            setor= 'SEXATAS'
        elif nomedepart == 'Departamento de Ciências do Solo e Engenharia Agrícola' or  nomedepart == 'Departamento de Engenharia Civil' or  nomedepart == 'Departamento de Engenharia de Alimentos' or  nomedepart == 'Departamento de Engenharia de Materiais' or  nomedepart == 'Departamento de Fitotecnia e Fitossanidade/DEFITO' or  nomedepart == 'Departamento de Informática' or  nomedepart == 'Departamento de Zootecnia':
            setor= 'SECATE'
        elif nomedepart == 'Departamento de Análises Clínicas' or   nomedepart == 'Departamento de Biologia Estrutural, Molecular e Genética' or  nomedepart == 'Departamento de Biologia Geral' or  nomedepart =='Departamento de Ciências Farmacêuticas' or  nomedepart == 'Departamento de Educação Física' or  nomedepart == 'Departamento de Enfermagem' or  nomedepart == 'Departamento de Medicina' or  nomedepart == 'Departamento de Odontologia' or  nomedepart == 'Departamento de Saúde Pública':
            setor = 'SEBISA'
        elif nomedepart == 'Departamento de Artes' or  nomedepart == 'Departamento de Educação' or  nomedepart == 'Departamento de História' or  nomedepart == 'Departamento de Estudos da Linguagem' or  nomedepart == 'Departamento de Pedagogia':
            setor = 'SECIHLA'
        elif nomedepart == 'Departamento de Direito das Relações Sociais' or  nomedepart == 'Departamento de Direito do Estado' or  nomedepart == 'Departamento de Direito Processual':
            setor = 'SECIJUR'
        elif nomedepart == 'Departamento de Administração' or  nomedepart == 'Departamento de Contabilidade' or  nomedepart == 'Departamento de Economia' or  nomedepart == 'Departamento de Jornalismo' or  nomedepart == 'Departamento de Serviço Social' or  nomedepart == 'Departamento de Turismo':
            setor = 'SECISA'
        # Executar a consulta
        def remover_acentos1(txt):
            if isinstance(txt, str):  # Verifica se o valor é uma string
                return ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')
            else:
                return ''  # Retorna uma string vazia se o valor não for uma string


        # Diretório das planilhas
        diretorio = "TIDE"

        # Inicializa a variável tide como 'nao'
        tide = 'nao'

        # Lista para armazenar os nomes das planilhas
        nomes_planilhas = []

        # Verifica se o diretório existe
        if os.path.exists(diretorio):
            # Percorre os arquivos no diretório
            for arquivo in os.listdir(diretorio):
                # Verifica se o arquivo é uma planilha
                if arquivo.endswith(".xlsx") or arquivo.endswith(".xls"):
                    # Lê a planilha
                    planilha = pd.read_excel(os.path.join(diretorio, arquivo), header=2)  # Começa a ler da terceira linha
                    # Remove acentos dos nomes na planilha
                    planilha['nm_pessoa_completo'] = planilha['nm_pessoa_completo'].apply(remover_acentos1)
                    # Adiciona os nomes das planilhas à lista
                    nomes_planilhas.extend(planilha['nm_pessoa_completo'].tolist())

        # Remove acentos do nome a ser procurado
        nome_sem_acentos = remover_acentos1(nome)

        # Verifica se o nome está presente nas planilhas
        if nome_sem_acentos in nomes_planilhas:
            tide = 'sim'
   
        cursor.execute("SELECT COUNT(*) FROM professores WHERE nome = %s", (nome,))
        total = cursor.fetchone()[0]

        # Verificar se o valor existe na tabela
        if total > 0:
            #print("O nome existe na tabela 'professores'."+(nome))
            cursor.execute("SELECT id_professor FROM professores WHERE nome= %s", (nome,))
            idprofessor = cursor.fetchone()[0]
        else:
            cursor.execute('''INSERT INTO public.professores (id_professor, nome, graduacao, departamento,status, setor, tide) 
                  VALUES (%s, %s, %s, %s, %s, %s, %s);''', 
                  (idprofessor, nome, varformacao, nomedepart, status, setor, tide))
            connection.commit()

        #         #INSERINDO PROJETOS NO BANCO
        def adiciona_projetos():
            anoprojeto = 0
            natureza = ''
            for participas in root.iter('ATIVIDADES-DE-PARTICIPACAO-EM-PROJETO'):
                part = participas.find('PARTICIPACAO-EM-PROJETO')

            for part in root.iter('PARTICIPACAO-EM-PROJETO'):
                pesquisa = part.find('PROJETO-DE-PESQUISA')

            for titulo in root.iter('PROJETO-DE-PESQUISA'):
                anoprojeto = titulo.attrib['ANO-INICIO']
                natureza = titulo.attrib['NATUREZA']
                tituloprojeto1 = titulo.attrib['NOME-DO-PROJETO']
                if tituloprojeto1 is not None:
                    tituloprojeto2 = unidecode.unidecode(tituloprojeto1) 
                    tituloprojeto3 = tituloprojeto2.upper()
                    tituloprojeto = tituloprojeto3.replace("'", "")
                cursor.execute("SELECT COUNT(*) FROM projetos WHERE nome = %s", (tituloprojeto,))
                tituloprojetocomparar = cursor.fetchone()[0]
                # Verificar se o valor existe na tabela
                if tituloprojetocomparar >0:
                    cursor.execute("SELECT id_professor FROM projetos WHERE nome = %s", (tituloprojeto,))
                    id_professorcomparar = cursor.fetchone()[0]
                    arrayid_professorcomparar = id_professorcomparar.split(",")
                    # Verificar se o valor existe na tabela
                    if idprofessor not in arrayid_professorcomparar :
                        cursor.execute("""      UPDATE projetos
                                                SET id_professor = CONCAT(id_professor, ',',%s)
                                                WHERE nome = %s;
                                            """,(idprofessor,tituloprojeto))
                        connection.commit()
                        cursor.execute("SELECT departamento FROM projetos WHERE nome= %s", (tituloprojeto,))
                        departcomparar = cursor.fetchone()[0]
                        arraydepartcomparar = departcomparar.split(',')
                        if nomedepart not in arraydepartcomparar:
                            #print("O nome do projeto existe na tabela 'projetos'."+(tituloprojeto))
                            cursor.execute("""  UPDATE projetos
                                                SET departamento = CONCAT(departamento, ',',%s) 
                                                WHERE nome = %s;
                                            """,(nomedepart, tituloprojeto))

                            connection.commit()
                else:
                    cursor.execute('''INSERT into public.projetos (id_professor,nome,anopubli,tipo,departamento) 
                    values('%s','%s','%s','%s','%s');
                        ''' % ((idprofessor),(tituloprojeto),(anoprojeto),(natureza),(nomedepart)))
                    connection.commit()

        #INSERINDO ARTIGOS NO BANCO
        def adiciona_artigos():
            anoartigo = 0
            tipoartigo = ''
            tituloartigo=''
            for artigos1 in root.iter('ARTIGOS-PUBLICADOS'):
                artigo1 = artigos1.find('ARTIGO-PUBLICADO')

            for artigos2 in root.iter('ARTIGO-PUBLICADO'):
                artigo2 = artigos2.find('DADOS-BASICOS-DO-ARTIGO')

            for artigo3 in root.iter('DADOS-BASICOS-DO-ARTIGO'):
                anoartigo = artigo3.attrib['ANO-DO-ARTIGO']
                idioma = artigo3.attrib['IDIOMA']
                tituloartigo1 = artigo3.attrib['TITULO-DO-ARTIGO']
                if tituloartigo1 is not None:
                    tituloartigo2 = unidecode.unidecode(tituloartigo1) 
                    tituloartigo3 = tituloartigo2.upper()
                    tituloartigo = tituloartigo3.replace("'", "")
                if idioma == 'Português':
                    tipoartigo='NACIONAL'
                else:
                    tipoartigo='INTERNACIONAL'

                cursor.execute("SELECT COUNT(*) FROM artigos WHERE nome = %s", (tituloartigo,))
                tituloartigocomparar = cursor.fetchone()[0]
                if tituloartigocomparar > 0:
                    cursor.execute("SELECT id_professor FROM artigos WHERE nome = %s", (tituloartigo,))
                    id_professor_atual = cursor.fetchone()[0]
                    arrayidprofessoratual = id_professor_atual.split(",")
                    if  idprofessor not in arrayidprofessoratual:
                        #print('prof diff')
                        cursor.execute("""      UPDATE artigos
                                                SET id_professor = CONCAT(id_professor, ',',%s)
                                                WHERE nome = %s;
                                            """,(idprofessor,tituloartigo))
                        connection.commit()
                        cursor.execute("SELECT departamento FROM artigos WHERE nome= %s", (tituloartigo,))
                        departcomparar = cursor.fetchone()[0]
                        if departcomparar is not None:
                            arraydepartcomparar = departcomparar.split(",")
                            print(arraydepartcomparar)
                            if nomedepart not in arraydepartcomparar :
                                #print('depart diff')
                                #print("O nome do artigo existe na tabela 'artigos'."+(tituloartigo))
                                cursor.execute("""  UPDATE artigos
                                                    SET departamento = CONCAT(departamento, ',', %s)
                                                    WHERE nome = %s;
                                                """,(nomedepart,tituloartigo))
                                connection.commit()
                
                else:
                    print('nao existe')
                    cursor.execute('''INSERT into public.artigos (id_professor, nome, anopubli, tipo, departamento) 
                    values (%s, %s, %s, %s, %s);
                    ''', (idprofessor, tituloartigo, anoartigo, tipoartigo, nomedepart))

                    connection.commit()
#          # ORIENTAÇÕES   
        def adiciona_ori():
            anoori = 0
            for prod in root.iter('OUTRA-PRODUCAO'):
                prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

            for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
                ori1 = ori.find('ORIENTACOES-CONCLUIDAS-PARA-MESTRADO')

            for ori1 in root.iter('ORIENTACOES-CONCLUIDAS-PARA-MESTRADO'):
                mestreano = ori1.find(
                    'DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO')
            for mestreano in root.iter('DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-MESTRADO'):
                anodomestrado = mestreano.attrib['ANO']
                tituloorimestrado1 = mestreano.attrib['TITULO']
                if tituloorimestrado1 is not None:
                    tituloorimestrado2 = unidecode.unidecode(tituloorimestrado1)
                    tituloorimestrado3 = tituloorimestrado2.upper()
                    tituloorimestrado = tituloorimestrado3.replace("'", "")
                    tipoori = 'MESTRADO'
                cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (tituloorimestrado,)),
                tituloorimestradocomparar = cursor.fetchone()[0]
                if tituloorimestradocomparar == 0:
                    cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                    values('%s','%s','%s','%s','%s');
                        ''' % ((idprofessor),(tituloorimestrado),(anodomestrado),(tipoori),(nomedepart)))
                    connection.commit()
            # DOUTORADO
            for prod in root.iter('OUTRA-PRODUCAO'):
                prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

            for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
                ori1 = ori.find('ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO')

            for ori1 in root.iter('ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO'):
                doutorano = ori1.find(
                    'DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO')
            for doutorano in root.iter('DADOS-BASICOS-DE-ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO'):
                anododoutorado = doutorano.attrib['ANO']
                titulooridoutorado1 = doutorano.attrib['TITULO']
                if titulooridoutorado1 is not None:
                    titulooridoutorado2 = unidecode.unidecode(titulooridoutorado1)
                    titulooridoutorado3 = titulooridoutorado2.upper()
                    titulooridoutorado = titulooridoutorado3.replace("'", "")
                    tipoori = 'DOUTORADO'
                cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (titulooridoutorado,))
                titulooridoutoradocomparar = cursor.fetchone()[0]
                # Verificar se o valor existe na tabela
                if titulooridoutoradocomparar == 0:
                    cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                    values('%s','%s','%s','%s','%s');
                        ''' % ((idprofessor),(titulooridoutorado),(anododoutorado),(tipoori),(nomedepart)))
                    connection.commit()

             # IC
            for prod in root.iter('OUTRA-PRODUCAO'):
                prod1 = prod.find('ORIENTACOES-CONCLUIDAS')

            for ori in root.iter('ORIENTACOES-CONCLUIDAS'):
                ori1 = ori.find('OUTRAS-ORIENTACOES-CONCLUIDAS')

            for ori1 in root.iter('OUTRAS-ORIENTACOES-CONCLUIDAS'):
                icano = ori1.find(
                    'DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS')
            for icano in root.iter('DADOS-BASICOS-DE-OUTRAS-ORIENTACOES-CONCLUIDAS'):
                anodoic = icano.attrib['ANO']
                deic = icano.attrib['NATUREZA']
                titulooriic1 = icano.attrib['TITULO']
                if deic == 'INICIACAO_CIENTIFICA':
                    if titulooriic1 is not None:
                        titulooriic2 = unidecode.unidecode(titulooriic1)
                        titulooriic3 = titulooriic2.upper()
                        titulooriic = titulooriic3.replace("'", "")
                        tipoori = 'INICIAÇÃO CIENTÍFICA'
                        cursor.execute("SELECT COUNT(*) FROM orientacoes WHERE nome = %s", (titulooriic,)),
                        titulooriiccomparar = cursor.fetchone()[0]
                        # Verificar se o valor existe na tabela
                        if titulooriiccomparar == 0:
                            cursor.execute('''INSERT into public.orientacoes (id_professor,nome,anoconclusao,tipo,departamento) 
                            values('%s','%s','%s','%s','%s');
                                ''' % ((idprofessor),(titulooriic),(anodoic),(tipoori),(nomedepart)))
                            connection.commit()


        # PUBLICACOES EM CONGRESSO
        def adiciona_congre():
            for biblio in root.iter('PRODUCAO-BIBLIOGRAFICA'):
                eventos = biblio.find('TRABALHOS-EM-EVENTOS')

            for eventos2 in root.iter('TRABALHOS-EM-EVENTOS'):
                evento = eventos2.find('TRABALHO-EM-EVENTOS')

            for evento2 in root.iter('TRABALHO-EM-EVENTOS'):
                dadoscongre = evento2.find('DADOS-BASICOS-DO-TRABALHO')
                tipocongre = dadoscongre.attrib['NATUREZA']
                anocongre = dadoscongre.attrib['ANO-DO-TRABALHO']
                titulocongre1 = dadoscongre.attrib['TITULO-DO-TRABALHO']
                if titulocongre1 is not None:
                    titulocongre2 = unidecode.unidecode(titulocongre1)
                    titulocongre3 = titulocongre2.upper()
                    titulocongre = titulocongre3.replace("'", "")
                    cursor.execute("SELECT COUNT(*) FROM congressos WHERE nome = %s", (titulocongre,))
                    titulocongrecomparar = cursor.fetchone()[0]
                    # Verificar se o valor existe na tabela
                    if titulocongrecomparar > 0:
                        cursor.execute("SELECT id_professor FROM congressos WHERE nome = %s", (titulocongre,))
                        id_professor_atual = cursor.fetchone()[0]
                        arrayidprofessoratual = id_professor_atual.split(",")
                        if idprofessor not in arrayidprofessoratual:
                            cursor.execute("""UPDATE congressos
                                            SET id_professor = CONCAT(id_professor, ',', %s)
                                            WHERE nome = %s""",
                                        (idprofessor, titulocongre))
                            connection.commit()
                            cursor.execute("SELECT departamento FROM congressos WHERE nome= %s", (titulocongre,))
                            departcomparar = cursor.fetchone()[0]
                            arraydepartcomparar = departcomparar.split(",")
                            if nomedepart not in arraydepartcomparar:
                                #print("O nome da publicaão existe na tabela 'congressos'." + titulocongre)
                                cursor.execute("""UPDATE congressos
                                                SET departamento = CONCAT(departamento, ',', %s)
                                                WHERE nome = %s""",
                                            (nomedepart, titulocongre))
                                connection.commit()

            
                    else:
                        cursor.execute('''INSERT into public.congressos (id_professor,nome,anoconclusao,tipo,departamento) 
                        values('%s','%s','%s','%s','%s');
                            ''' % ((idprofessor),(titulocongre),(anocongre),(tipocongre),(nomedepart)))
                        connection.commit()
        
        # Função para comparar os nomes ignorando os acentos
        #print('ORIENTACOES')
        adiciona_ori()
        #print('ARTIGOS')
        adiciona_artigos()

        #print('CONGRESSOS')
        adiciona_congre()

        #print('PROJETOS')
        adiciona_projetos()

        def comparar_nomes(nome_a, nome_b):
            return unidecode.unidecode(str(nome_a).lower()) == unidecode.unidecode(str(nome_b).lower())

        # Leitura das planilhas
        colaboradores_df = pd.read_excel("COLABORADORES_SEXATAS.xlsx")
        efetivos_df = pd.read_excel("Efetivos_SEXATAS.xlsx")
        tide_df = pd.read_excel("TIDE_SEXATAS.xlsx")

        # Extrair os valores da coluna C, ignorando os valores que não são strings
        colaboradores_nomes = [str(nome) for nome in colaboradores_df.iloc[:, 2].tolist() if isinstance(nome, str)]
        efetivos_nomes = [str(nome) for nome in efetivos_df.iloc[:, 2].tolist() if isinstance(nome, str)]
        tide_nomes = [str(nome) for nome in tide_df.iloc[:, 2].tolist() if isinstance(nome, str)]


        # def percorrer_planilhas():
        #     anosUEPG = []  # Inicialize a lista de anosUEPG
        #     for filename in os.listdir("PLANILHAS"):
        #         if filename.endswith(".xlsx"):
        #             path = os.path.join("PLANILHAS", filename)
        #             df = pd.read_excel(path)
        #             nomes = [str(nome) for nome in df.iloc[:, 2].tolist() if isinstance(nome, str)]
        #             anos = df.iloc[:, 0].tolist()
        #             for i, n in enumerate(nomes):
        #                 if comparar_nomes(nome, n):
        #                     ano = anos[i]
        #                     if isinstance(ano, str) and ano.strip() != '' and ano.lower() != 'nan':
        #                         anosUEPG.append(str(ano))
        #     return ','.join(anosUEPG)
        # Chamada da função para percorrer as planilhas
        #anosUEPG = percorrer_planilhas()
        #print("Anos da UEPG para", nome, ":", anosUEPG)
        def remover_acentos(txt):
            if txt is None:
                return ""
            return ''.join(ch for ch in unicodedata.normalize('NFD', str(txt)) if not unicodedata.combining(ch))

        def ler_planilhas(pasta, nome):
            anosUEPG = []
            for arquivo in os.listdir(pasta):
                if arquivo.endswith('.xlsx'):
                    caminho_arquivo = os.path.join(pasta, arquivo)
                    planilha = load_workbook(caminho_arquivo)
                    for nome_planilha in planilha.sheetnames:
                        planilha_ativa = planilha[nome_planilha]
                        for linha in planilha_ativa.iter_rows(min_row=3, min_col=1, max_col=3, values_only=True):
                            ano, _, nome_planilha = linha
                            nome_sem_acentos = remover_acentos(nome_planilha).lower()
                            if nome_sem_acentos == nome.lower():
                                if ano not in anosUEPG:
                                    anosUEPG.append(str(ano))
            return anosUEPG

        # Diretório do projeto
        diretorio_projeto = os.path.dirname(os.path.abspath(__file__))

        # Defina as pastas onde estão as planilhas
        pasta_colaboradores = os.path.join(diretorio_projeto, 'COLABORADORES')
        pasta_efetivos = os.path.join(diretorio_projeto, 'EFETIVOS')

        anos_colaboradores = ler_planilhas(pasta_colaboradores, nome)
        anos_efetivos = ler_planilhas(pasta_efetivos, nome)

        anosUEPG = anos_colaboradores + anos_efetivos
        anosUEPG_concatenados = ', '.join(anosUEPG)

        #print(f"Anos da UEPG para '{nome}': {anosUEPG_concatenados}")
        anosUEPG_unicos = list(set(anosUEPG))
        # Ordenando a lista para manter os anos em ordem crescente
        anosUEPG_unicos.sort()
        cursor.execute("""  
            UPDATE professores
            SET anosUEPG = %s
            WHERE nome = %s;
        """, (anosUEPG_unicos, nome))
        connection.commit()

        #PLANILHAS BOLSAs

        # Obtém o diretório atual onde está o código
        diretorio_atual = os.path.dirname(os.path.abspath(__file__))

        # # Pasta onde estão os dados
        pasta_bolsasic = "BOLSASIC"
        # Percorre os anos dentro da pasta BOLSASIC
        for ano in range(2019, 2025):  # Modifique conforme os anos desejados
            pasta_ano = os.path.join(diretorio_atual, pasta_bolsasic, str(ano))

            if os.path.exists(pasta_ano):
                # Lista os arquivos na pasta do ano
                arquivos = os.listdir(pasta_ano)
                # Itera sobre os arquivos
                for arquivo in arquivos:
                    # Verifica se o arquivo é uma planilha (pode adaptar essa verificação conforme o tipo de arquivo)
                    if arquivo.endswith('.xlsx') or arquivo.endswith('.xls'):
                        # Lê a planilha
                        caminho_arquivo = os.path.join(pasta_ano, arquivo)
                        
                        # Lê a planilha, especificando que o cabeçalho está na linha 8 (0-indexed)
                        df = pd.read_excel(caminho_arquivo, header=8)
                        df['ORIENTADOR(A)'] = df['ORIENTADOR(A)'].astype(str).apply(unidecode.unidecode)
                        df['SUBPROJETO DO'] = df['SUBPROJETO DO'].astype(str).apply(unidecode.unidecode)
                  
                        # Filtra as linhas onde o nome do professor corresponde ao nome desejado
                        df_filtrado = df[df['ORIENTADOR(A)'] == unidecode.unidecode(nome)]
                        
                        # Itera sobre as linhas filtradas

                        for index, row in df_filtrado.iterrows():
                            # Imprime o nome do trabalho e o ano da pasta
                            titulobolsa = row['SUBPROJETO DO']
                            anobolsa = ano
                             # Adiciona uma linha em branco entre as impressões

                            #Verifica se o título da bolsa já existe na tabela

                            # Executar consulta somente se a conexão estiver estabelecida corretamente
                            if connection:
                                cursor.execute("SELECT COUNT(*) FROM bolsas WHERE nome = %s", (titulobolsa,))
                                titulobolsacomparar = cursor.fetchone()[0]
                            # Se o título da bolsa não existir na tabela, insere os dados
                            if titulobolsacomparar < 1:

                                cursor.execute('''INSERT INTO public.bolsas (id_professor, nome, ano, departamento) 
                                                VALUES (%s, %s, %s, %s);''',
                                            (idprofessor, titulobolsa, anobolsa, nomedepart))
                                connection.commit()

        #FINANCIADOS
        def remover_acentosfinan(texto):
            return unidecode.unidecode(str(texto)) if texto else texto

        def encontrar_docente(nome):
            pasta = "FINANCIADOS"
            # Lista todos os arquivos na pasta especificada
            arquivos = os.listdir(pasta)
            
            # Itera sobre cada arquivo na pasta
            for arquivo in arquivos:
                # Verifica se o arquivo é uma planilha Excel
                if arquivo.endswith('.xlsx'):
                    # Carrega a planilha
                    wb = load_workbook(os.path.join(pasta, arquivo))
                    # Acessa a primeira planilha (índice 0)
                    planilha = wb.worksheets[0]
                    
                    # Itera sobre as células das colunas I (nome do docente), B (nome do projeto) e G (valor do projeto) a partir da linha 3
                    for docente, projeto, valor, anofinanciado_celula in zip(planilha['I'][2:], planilha['B'][2:], planilha['G'][2:], planilha['J'][2:]):
                        nome_docente = remover_acentosfinan(docente.value)
                        
                        #print (nome,nome_docente)
                        if nome_docente == nome:
                            nome_projeto = remover_acentosfinan(projeto.value)
                            valor_projeto = remover_acentosfinan(valor.value)
                            anofinanciado_str = str(anofinanciado_celula.value)
                    
                            # Verificar se a célula não está vazia
                
                            # Verificar se a célula não está vazia e não é "None"
                            if anofinanciado_str and anofinanciado_str.strip():
                                anofinanciado_datetime = datetime.strptime(anofinanciado_str, "%Y-%m-%d %H:%M:%S")
                                # Extrair o ano
                                anofinanciado = anofinanciado_datetime.year
                            else:
                                # Se a célula estiver vazia ou for None, defina anofinanciado como None ou outro valor padrão
                                anofinanciado = None
                            # Remova o "R$" e quaisquer caracteres não numéricos do valor do projeto
                            valor_projeto = valor_projeto.replace('R$', '').replace(',', '').strip()
                            # Converte o valor para float
                            valor_projeto = float(valor_projeto)
                            cursor.execute("SELECT COUNT(*) FROM financiados WHERE nome = %s", (nome_projeto,))
                            titulofinanciadocomparar = cursor.fetchone()[0]
                            # Verificar se o valor existe na tabela
                            if titulofinanciadocomparar == 0:
                                cursor.execute('''INSERT into public.financiados (id_professor,nome,anopubli,valor,departamento) 
                                                values('%s','%s','%s','%f','%s');
                                                ''' % ((idprofessor),(nome_projeto),(anofinanciado),(valor_projeto),(nomedepart)))
                                connection.commit()
        # Chama a função para buscar o docente na pasta "FINANCIADOS"
        encontrar_docente(nome)

        
        # Diretório onde estão os arquivos .xlsx
        diretorio = "PRODUTIVIDADE"

        # Verifica se o diretório existe
        if not os.path.exists(diretorio):
            #print(f"O diretório {diretorio} não existe.")
            exit()

        # Verifica se o diretório está vazio
        if not os.listdir(diretorio):
            #print(f"O diretório {diretorio} está vazio.")
            exit()

        # Função para gerar array de anos
        def gerar_anos(inicio, termino):
            anos = []
            ano_inicio = inicio.year
            ano_termino = termino.year
            for ano in range(ano_inicio, ano_termino + 1):
                anos.append(str(ano))
            return anos

        # Inicializa a lista de anos de pesquisa
        anos_pesquisa = []

        # Conexão com o banco de dados
        def inserir_dados(id_professor, ano, departamento):
                # Verifica se já existe uma linha com o mesmo id_professor e ano
                cursor.execute('''SELECT 1 FROM public.produtividade 
                                WHERE id_professor = %s AND ano = %s;''', (id_professor, ano))
                row = cursor.fetchone()
                if row:
                    print(f"Já existe uma linha para o professor {id_professor} no ano {ano}.")
                else:
                    # Insere os dados apenas se não houver linha duplicada
                    cursor.execute('''INSERT INTO public.produtividade (id_professor, ano, departamento) 
                                    VALUES (%s, %s, %s);''', (id_professor, ano, departamento))
                    #print(f"Dados inseridos para o professor {id_professor} no ano {ano}.")
                connection.commit()



        # Percorre todos os arquivos .xlsx no diretório especificado
        for arquivo in os.listdir(diretorio):
            if arquivo.endswith(".xlsx"):
                # Carrega o arquivo Excel
                caminho_arquivo = os.path.join(diretorio, arquivo)
                df = pd.read_excel(caminho_arquivo)

                # Remove acentos dos nomes dos pesquisadores
                df['PESQUISADOR'] = df['PESQUISADOR'].apply(lambda x: unidecode.unidecode(str(x)))

                # Filtra as linhas onde o nome do pesquisador corresponde
                df_pesquisador = df[df['PESQUISADOR'] == unidecode.unidecode(nome)]

                # Para cada linha do pesquisador, calcula os anos de pesquisa
                for index, row in df_pesquisador.iterrows():
                    data_inicio = pd.to_datetime(row['INÍCIO'], format='%d/%m/%Y')
                    data_termino = pd.to_datetime(row['TÉRMINO'], format='%d/%m/%Y')
                    anos = gerar_anos(data_inicio, data_termino)
                    anos_pesquisa.extend(anos)

                    # Inserir os dados no banco para cada ano
                    for ano in anos:
                        inserir_dados(idprofessor, ano, departamento)

        # Remove anos duplicados
        anos_pesquisa = list(set(anos_pesquisa))
        #print(anos_pesquisa)


except (Exception, psycopg2.Error) as error:
    print("Erro ao conectar ao PostgreSQL:", error)
finally:
# Fechar a conexão
    if connection:
        cursor.close()
        connection.close()